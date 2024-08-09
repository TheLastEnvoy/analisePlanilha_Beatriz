import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Função para carregar e processar as planilhas
def carregar_planilhas():
    st.sidebar.header("Carregar Planilhas")
    upload_planilha_ontem = st.sidebar.file_uploader("Selecionar Planilha de Ontem", type=["xlsx"])
    upload_planilha_hoje = st.sidebar.file_uploader("Selecionar Planilha de Hoje", type=["xlsx"])

    def processar_planilha(upload):
        if upload is not None:
            # Ler todas as abas e concatenar em um único DataFrame, ignorando a primeira aba
            xls = pd.read_excel(upload, sheet_name=None, engine='openpyxl')
            sheets = []
            for idx, (name, df) in enumerate(xls.items()):
                if idx == 0:  # Ignorar a primeira aba
                    continue
                # Procurar a linha que contém os cabeçalhos específicos
                header_row_index = None
                for i in range(len(df)):
                    if 'Lote' in df.iloc[i].values and 'Carga' in df.iloc[i].values and 'Produto' in df.iloc[i].values:
                        header_row_index = i
                        break
                if header_row_index is None:
                    st.error(f"Os cabeçalhos esperados não foram encontrados na aba '{name}'.")
                    return None

                df.columns = df.iloc[header_row_index]  # Usa a linha do cabeçalho real como cabeçalho
                df = df[(header_row_index + 1):]  # Remove as linhas acima do cabeçalho real
                df = df.reset_index(drop=True)  # Resetar os índices
                sheets.append(df)
            if sheets:
                return pd.concat(sheets, ignore_index=True)
            else:
                st.error("Nenhuma aba válida foi encontrada nas planilhas.")
                return None
        return None

    planilha_anterior = processar_planilha(upload_planilha_ontem)
    planilha_atual = processar_planilha(upload_planilha_hoje)

    return planilha_anterior, planilha_atual

# Função para ajustar a largura das colunas
def ajustar_largura_colunas(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Coluna A, B, C, etc.
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[col_letter].width = adjusted_width

# Função para executar o código principal
def executar_codigo(planilha_anterior, planilha_atual):
    try:
        # Considerando que a coluna chave para identificar os lotes seja 'Lote'
        coluna_chave = 'Lote'

        # Verificar se a coluna 'Lote' existe em ambas as planilhas
        if coluna_chave not in planilha_anterior.columns or coluna_chave not in planilha_atual.columns:
            raise KeyError(f"A coluna '{coluna_chave}' não foi encontrada em uma ou ambas as planilhas.")

        # Encontrar os lotes novos que estão na planilha atual, mas não estavam na planilha anterior
        novos_lotes = planilha_atual[~planilha_atual[coluna_chave].isin(planilha_anterior[coluna_chave])]

        # Criar uma nova planilha Excel
        nome_arquivo_saida = 'novos_lotes_identificados.xlsx'
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Planilha'

        # Adicionar cabeçalho da planilha
        header = list(planilha_atual.columns)
        sheet.append(header)

        # Copiar dados da planilha_atual para a nova planilha
        for r in dataframe_to_rows(planilha_atual, index=False, header=False):
            sheet.append(r)

        # Aplicar formatação às linhas dos novos itens (linhas marcadas como novos)
        gray_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column), start=2):
            if planilha_atual.loc[row_idx - 2, 'Lote'] in novos_lotes['Lote'].values:
                for cell in row:
                    cell.fill = gray_fill

        # Ajustar a largura das colunas
        ajustar_largura_colunas(sheet)

        # Salvar a planilha em um objeto binário
        from io import BytesIO
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        output.seek(0)

        st.success(f"Foram identificados {len(novos_lotes)} novos lotes.")

        return output, nome_arquivo_saida

    except KeyError as e:
        st.error(str(e))
        return None, None
    except Exception as e:
        st.error(f"Ocorreu um erro ao executar o código:\n{str(e)}")
        return None, None

# Título do dashboard
st.title("Ferramenta para análise de novos lotes")

# Carregar planilhas
planilha_anterior, planilha_atual = carregar_planilhas()

# Botão para executar o código
if st.sidebar.button("Executar Código"):
    if planilha_anterior is not None and planilha_atual is not None:
        output, nome_arquivo_saida = executar_codigo(planilha_anterior, planilha_atual)
        if output is not None:
            st.download_button(
                label="Download da planilha gerada",
                data=output,
                file_name=nome_arquivo_saida,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning("Por favor, selecione ambas as planilhas antes de executar.")
